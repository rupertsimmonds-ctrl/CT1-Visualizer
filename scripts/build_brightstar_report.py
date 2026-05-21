"""Build the Brightstar portfolio report as an xlsx workbook.

Source data: 18 line items pasted by Rupert from Asset_inventory.xlsx.
Unit and key counts: researched from public sources, cited in Notes & Sources.
Where no public source gave a firm count, the cell is left blank rather than guessed.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# betterhomes brand palette (from CT1-Visualizer :root)
SLATE = "1F343F"
DENIM = "2C537A"
POWDER = "7BA0B2"
SAND = "D9B9A0"
MIST = "EDE8E4"
BRONZE = "B39470"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1 = Font(name="Calibri", size=18, bold=True, color=WHITE)
H2 = Font(name="Calibri", size=12, bold=True, color=WHITE)
H3 = Font(name="Calibri", size=11, bold=True, color=SLATE)
BODY = Font(name="Calibri", size=11, color=SLATE)
BODY_MUTED = Font(name="Calibri", size=10, color="5A5A5A")
BODY_BOLD = Font(name="Calibri", size=11, bold=True, color=SLATE)

FILL_SLATE = PatternFill("solid", fgColor=SLATE)
FILL_DENIM = PatternFill("solid", fgColor=DENIM)
FILL_POWDER = PatternFill("solid", fgColor=POWDER)
FILL_SAND = PatternFill("solid", fgColor=SAND)
FILL_MIST = PatternFill("solid", fgColor=MIST)
FILL_BRONZE = PatternFill("solid", fgColor=BRONZE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# Inventory: 18 line items from Rupert's paste, consolidated to physical assets.
# Columns: line item, physical asset, segment, sub-use, GLA, count metric,
# count value, count source key.
# ---------------------------------------------------------------------------

INVENTORY = [
    # line item,                 physical asset,          segment,      sub-use,        GLA,     metric,         count,  src
    ("Falcon Tower - RT",        "Falcon Tower",          "Retail",     "Ground retail",  21850, "retail units", None,   "s_falcon"),
    ("Falcon Tower - RE",        "Falcon Tower",          "Residential","Apartments",   182400, "apartments",   None,   "s_falcon"),
    ("City Tower 2 - RT",        "City Tower 2",          "Retail",     "Podium retail",  16664, "retail units", None,   "s_ct2"),
    ("City Tower 2 - O",         "City Tower 2",          "Office",     "Offices",      255582, "office floors", 22,    "s_ct2"),
    ("Green Tower",              "Green Tower",           "Office",     "Offices",      120766, "office floors", 17,    "s_green"),
    ("Splendour Villas",         "Splendour Villas",      "Residential","Villas",       357512, "villas",         90,    "s_splendour"),
    ("Al Sufouh Villas",         "Al Sufouh Villas",      "Residential","Villas",       111936, "villas",         32,    "s_sufouh"),
    ("Galleria Villas",          "Galleria Villas Al Wasl","Residential","Villas",      170852, "villas",         38,    "s_galleria_v"),
    ("Al Barsha Villas - A",     "Al Barsha Villas A",    "Residential","Villas",        65388, "villas",        None,   "s_barsha_v"),
    ("Al Barsha Villas - B",     "Al Barsha Villas B",    "Residential","Villas",        27986, "villas",        None,   "s_barsha_v"),
    ("Al Barsha Villas - C",     "Al Barsha Villas C",    "Residential","Villas",        45550, "villas",        None,   "s_barsha_v"),
    ("EH - Jumeirah Garden",     "EH Jumeirah Garden",    "Residential","Apartments",   280277, "apartments",    None,   "s_eh"),
    ("Galleria Al Wasl",         "Galleria Al Wasl",      "Retail",     "Boutique mall", 98928, "retail units",   27,    "s_gal_wasl"),
    ("Galleria Al Barsha",       "Galleria Al Barsha",    "Retail",     "Boutique mall", 81862, "retail units",  None,   "s_gal_barsha"),
    ("Millenium Tower",          "Millennium Tower",      "Residential","Apartments",   614307, "apartments",    407,   "s_millennium"),
    ("Medcare Hospital",         "Medcare Hospital",      "Hospital",   "Hospital",     108204, "beds",           64,   "s_medcare"),
    ("Four Seasons Jumeirah",    "Four Seasons Jumeirah Beach","Hotel", "Resort",       674000, "keys",          237,   "s_fs_jb"),
    ("Four Seasons DFC",         "Four Seasons DIFC",     "Hotel",      "City hotel",   245000, "keys",          106,   "s_fs_difc"),
]

SOURCES = {
    "s_falcon":      ("Falcon Tower, Business Bay: 41 floors, 1 to 3 bed apartments + ground retail. No public total unit count.", "https://www.bayut.com/buildings/falcon-tower-business-bay/"),
    "s_ct2":         ("City Tower 2, Sheikh Zayed Road: 22-storey office building completed 1997.", "https://www.bayut.com/buildings/city-tower-2/"),
    "s_green":       ("Green Tower, Riggat Al Buteen, Deira: 17-storey office building completed 1999.", "https://propsearch.ae/dubai/green-tower"),
    "s_splendour":   ("Splendour Villas, Al Safa 1: 90 four-bed duplex villas, gated, H&H managed.", "https://www.h-hpm.ae/developments/splendor-villas/"),
    "s_sufouh":      ("Sufouh Villas: 32 four-bed villas + clubhouse, H&H managed.", "https://www.h-hpm.ae/developments/sufouh-villas/"),
    "s_galleria_v":  ("The Galleria Villas, Al Wasl: 38 four to five bed villas with private pools, H&H managed.", "https://www.h-hpm.ae/developments/the-galleria-villas-al-wasl/"),
    "s_barsha_v":    ("Al Barsha Villas A, B, C: three compounds in Al Barsha. No public unit count by sub-compound found.", ""),
    "s_eh":          ("EH Jumeirah Garden: likely residential block in Jumeirah Garden City, Al Satwa. No public unit count under this exact name.", ""),
    "s_gal_wasl":    ("Galleria Mall Al Wasl: anchor + 22 retail + 4 F&B over two floors plus rooftop open-air cinema. Hopkins Architects cites 34 boutique offers.", "https://www.hopkins.co.uk/projects/hospitality/galleria-mall-jumeirah/"),
    "s_gal_barsha":  ("Galleria Mall Al Barsha: two-storey boutique retail centre, c.17,493 sqm built area. No firm public retail-unit count.", "https://uaemood.com/galleria-mall-al-barsha-dubai/"),
    "s_millennium":  ("Millennium Tower, Sheikh Zayed Road: 60 floors, 301 three-bed + 106 two-bed = 407 apartments.", "https://en.wikipedia.org/wiki/Millennium_Tower_(Dubai)"),
    "s_medcare":     ("Medcare Hospital, Al Safa, Jumeirah: 64-bed JCI-accredited multi-specialty hospital.", "https://www.medcare.ae/en/branches/view/medcare-hospital.html"),
    "s_fs_jb":       ("Four Seasons Resort Dubai at Jumeirah Beach: 237 keys, including 49 suites.", "https://press.fourseasons.com/dubaijb/hotel-facts/"),
    "s_fs_difc":     ("Four Seasons Hotel DIFC: 106 keys (78 rooms + 28 suites).", "https://www.fourseasons.com/dubaidifc/"),
}

SEGMENT_ORDER = ["Residential", "Hotel", "Office", "Retail", "Hospital"]

# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def title_band(ws, row, text, span, fill, font, height=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height


def header_row(ws, row, headers, fill=FILL_DENIM, font=H2):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

wb = Workbook()

# Aggregations
total_gla = sum(r[4] for r in INVENTORY)
gla_by_segment = {s: 0 for s in SEGMENT_ORDER}
line_count_by_segment = {s: 0 for s in SEGMENT_ORDER}
for r in INVENTORY:
    gla_by_segment[r[2]] += r[4]
    line_count_by_segment[r[2]] += 1

physical_assets = sorted({r[1] for r in INVENTORY})
physical_by_segment = {}
for r in INVENTORY:
    physical_by_segment.setdefault(r[2], set()).add(r[1])

# Group physical assets into a category (towers / villas / retail / hotels / hospital / other residential)
def category_of(asset_name, segment):
    name = asset_name.lower()
    if segment == "Hotel": return "Hotels"
    if segment == "Hospital": return "Hospital"
    if "tower" in name: return "Towers"
    if "villa" in name: return "Villa communities"
    if "galleria" in name and segment == "Retail": return "Retail centres"
    if segment == "Residential": return "Other residential"
    if segment == "Retail": return "Retail centres"
    if segment == "Office": return "Towers"
    return "Other"

category_assets = {}
for r in INVENTORY:
    cat = category_of(r[1], r[2])
    category_assets.setdefault(cat, set()).add(r[1])

# ===========================================================================
# Sheet 1: Summary
# ===========================================================================
ws = wb.active
ws.title = "Summary"
set_widths(ws, [34, 18, 18, 18, 18, 18])

title_band(ws, 1, "Brightstar portfolio report", 6, FILL_SLATE, H1, height=36)
title_band(ws, 2, "Source: Asset_inventory.xlsx (18 line items). Owner assumed to be Brightstar; not stated on file.", 6, FILL_MIST, BODY_MUTED, height=20)

# Headline tiles
ws.cell(row=4, column=1, value="Headline").font = H3
row = 5
tiles = [
    ("Line items in file", "18"),
    ("Physical assets", str(len(physical_assets))),
    ("Total GLA (sqf)", f"{total_gla:,}"),
    ("Segments", str(len([s for s in SEGMENT_ORDER if gla_by_segment[s] > 0]))),
]
for i, (label, value) in enumerate(tiles):
    col = 1 + i * 2
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = FILL_POWDER; lc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    lc.alignment = CENTER; lc.border = BORDER
    vc = ws.cell(row=row+1, column=col, value=value)
    vc.fill = FILL_MIST; vc.font = Font(name="Calibri", size=16, bold=True, color=SLATE)
    vc.alignment = CENTER; vc.border = BORDER
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
ws.row_dimensions[row].height = 22
ws.row_dimensions[row+1].height = 34

# GLA by segment
row = 9
ws.cell(row=row, column=1, value="GLA by segment").font = H3
row += 1
header_row(ws, row, ["Segment", "Line items", "Physical assets", "GLA (sqf)", "Share of GLA", ""])
row += 1
for seg in SEGMENT_ORDER:
    if gla_by_segment[seg] == 0:
        continue
    ws.cell(row=row, column=1, value=seg).font = BODY_BOLD
    ws.cell(row=row, column=2, value=line_count_by_segment[seg]).alignment = CENTER
    ws.cell(row=row, column=3, value=len(physical_by_segment.get(seg, set()))).alignment = CENTER
    c = ws.cell(row=row, column=4, value=gla_by_segment[seg]); c.number_format = "#,##0"; c.alignment = RIGHT
    c = ws.cell(row=row, column=5, value=gla_by_segment[seg] / total_gla); c.number_format = "0.0%"; c.alignment = RIGHT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = BORDER
    row += 1
# total
ws.cell(row=row, column=1, value="Total").font = BODY_BOLD
ws.cell(row=row, column=1).fill = FILL_SAND
ws.cell(row=row, column=2, value=len(INVENTORY)).alignment = CENTER; ws.cell(row=row, column=2).fill = FILL_SAND
ws.cell(row=row, column=3, value=len(physical_assets)).alignment = CENTER; ws.cell(row=row, column=3).fill = FILL_SAND
c = ws.cell(row=row, column=4, value=total_gla); c.number_format = "#,##0"; c.alignment = RIGHT; c.fill = FILL_SAND
c = ws.cell(row=row, column=5, value=1.0); c.number_format = "0.0%"; c.alignment = RIGHT; c.fill = FILL_SAND
for col in range(1, 6):
    ws.cell(row=row, column=col).border = BORDER
    ws.cell(row=row, column=col).font = BODY_BOLD

# Asset mix by category
row += 3
ws.cell(row=row, column=1, value="Asset mix by category").font = H3
row += 1
header_row(ws, row, ["Category", "Physical assets", "Names", "", "", ""])
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
row += 1
cat_order = ["Towers", "Villa communities", "Retail centres", "Hotels", "Hospital", "Other residential"]
for cat in cat_order:
    names = sorted(category_assets.get(cat, []))
    if not names:
        continue
    ws.cell(row=row, column=1, value=cat).font = BODY_BOLD
    ws.cell(row=row, column=2, value=len(names)).alignment = CENTER
    nc = ws.cell(row=row, column=3, value=", ".join(names))
    nc.alignment = LEFT
    nc.font = BODY
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 30
    row += 1

# Researched counts callout
row += 2
ws.cell(row=row, column=1, value="Unit and key counts (researched)").font = H3
row += 1
header_row(ws, row, ["Metric", "Count", "Source", "", "", ""])
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
row += 1
researched = [
    ("Hotel keys (Four Seasons Jumeirah Beach + DIFC)", 237 + 106, "Four Seasons press kits"),
    ("Hospital beds (Medcare)", 64, "Medcare website"),
    ("Apartments (Millennium Tower)", 407, "Wikipedia, building data"),
    ("Villas (Splendour + Sufouh + Galleria Al Wasl)", 90 + 32 + 38, "H&H Property Management portfolio"),
]
for label, val, src in researched:
    ws.cell(row=row, column=1, value=label).font = BODY
    c = ws.cell(row=row, column=2, value=val); c.number_format = "#,##0"; c.alignment = CENTER; c.font = BODY_BOLD
    sc = ws.cell(row=row, column=3, value=src); sc.font = BODY_MUTED
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = BORDER
    row += 1

# Caveat band
row += 2
title_band(ws, row, "Caveats: source file has no unit counts. Counts shown above are from public sources (cited on Notes sheet). Blank counts mean no public figure was confirmed; do not guess.", 6, FILL_BRONZE, Font(name="Calibri", size=10, bold=True, color=WHITE), height=32)

ws.freeze_panes = "A4"

# ===========================================================================
# Sheet 2: Assets (full line-item inventory)
# ===========================================================================
ws = wb.create_sheet("Assets")
set_widths(ws, [26, 28, 14, 18, 14, 14, 12, 38])

title_band(ws, 1, "Asset inventory, 18 line items, consolidated to physical asset", 8, FILL_SLATE, H1, height=32)
header_row(ws, 2, [
    "Line item (as in file)",
    "Physical asset",
    "Segment",
    "Sub-use",
    "GLA (sqf)",
    "Count metric",
    "Count",
    "Source / note",
])

row = 3
for line, asset, seg, sub, gla, metric, count, src_key in INVENTORY:
    ws.cell(row=row, column=1, value=line).font = BODY
    ws.cell(row=row, column=2, value=asset).font = BODY_BOLD
    ws.cell(row=row, column=3, value=seg).alignment = CENTER
    ws.cell(row=row, column=4, value=sub).font = BODY
    c = ws.cell(row=row, column=5, value=gla); c.number_format = "#,##0"; c.alignment = RIGHT
    ws.cell(row=row, column=6, value=metric).alignment = CENTER
    cc = ws.cell(row=row, column=7, value=count if count is not None else "n/a")
    cc.alignment = CENTER
    cc.font = BODY_BOLD if count is not None else BODY_MUTED
    src_text, src_url = SOURCES[src_key]
    sc = ws.cell(row=row, column=8, value=src_text)
    sc.font = BODY_MUTED
    sc.alignment = LEFT
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = BORDER
    # stripe
    if row % 2 == 1:
        for col in range(1, 9):
            if ws.cell(row=row, column=col).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=row, column=col).fill = FILL_MIST
    ws.row_dimensions[row].height = 32
    row += 1

# total row
ws.cell(row=row, column=1, value="Total").font = BODY_BOLD
ws.cell(row=row, column=1).fill = FILL_SAND
c = ws.cell(row=row, column=5, value=total_gla); c.number_format = "#,##0"; c.alignment = RIGHT; c.font = BODY_BOLD; c.fill = FILL_SAND
for col in (2, 3, 4, 6, 7, 8):
    ws.cell(row=row, column=col).fill = FILL_SAND
for col in range(1, 9):
    ws.cell(row=row, column=col).border = BORDER

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{row-1}"

# ===========================================================================
# Sheet 3: By segment
# ===========================================================================
ws = wb.create_sheet("Segments")
set_widths(ws, [22, 14, 18, 18, 14, 38])

title_band(ws, 1, "Portfolio by segment", 6, FILL_SLATE, H1, height=32)
header_row(ws, 2, ["Segment", "Line items", "Physical assets", "GLA (sqf)", "Share", "Assets"])

row = 3
for seg in SEGMENT_ORDER:
    if gla_by_segment[seg] == 0:
        continue
    assets = sorted(physical_by_segment.get(seg, set()))
    ws.cell(row=row, column=1, value=seg).font = BODY_BOLD
    ws.cell(row=row, column=2, value=line_count_by_segment[seg]).alignment = CENTER
    ws.cell(row=row, column=3, value=len(assets)).alignment = CENTER
    c = ws.cell(row=row, column=4, value=gla_by_segment[seg]); c.number_format = "#,##0"; c.alignment = RIGHT
    c = ws.cell(row=row, column=5, value=gla_by_segment[seg] / total_gla); c.number_format = "0.0%"; c.alignment = RIGHT
    ac = ws.cell(row=row, column=6, value=", ".join(assets)); ac.font = BODY; ac.alignment = LEFT
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 36
    row += 1

# total
ws.cell(row=row, column=1, value="Total").font = BODY_BOLD; ws.cell(row=row, column=1).fill = FILL_SAND
ws.cell(row=row, column=2, value=len(INVENTORY)).alignment = CENTER; ws.cell(row=row, column=2).fill = FILL_SAND
ws.cell(row=row, column=3, value=len(physical_assets)).alignment = CENTER; ws.cell(row=row, column=3).fill = FILL_SAND
c = ws.cell(row=row, column=4, value=total_gla); c.number_format = "#,##0"; c.alignment = RIGHT; c.fill = FILL_SAND; c.font = BODY_BOLD
c = ws.cell(row=row, column=5, value=1.0); c.number_format = "0.0%"; c.alignment = RIGHT; c.fill = FILL_SAND; c.font = BODY_BOLD
ws.cell(row=row, column=6).fill = FILL_SAND
for col in range(1, 7):
    ws.cell(row=row, column=col).border = BORDER

ws.freeze_panes = "A3"

# ===========================================================================
# Sheet 4: Notes & Sources
# ===========================================================================
ws = wb.create_sheet("Notes & Sources")
set_widths(ws, [28, 60, 50])

title_band(ws, 1, "Notes, assumptions, and sources", 3, FILL_SLATE, H1, height=32)

row = 3
ws.cell(row=row, column=1, value="Assumptions").font = H3
row += 1
notes = [
    ("Ownership", "The source file does not name an owner. Treated as Brightstar-owned. Confirm before external use."),
    ("Consolidation", "Falcon Tower (Retail + Residential), City Tower 2 (Retail + Office), and Al Barsha Villas (A + B + C) are split across line items but are one physical asset each. Report shows both line-item count (18) and physical-asset count (15)."),
    ("Unit counts", "Source file has no unit counts. Counts shown were researched from public listings and operator websites and are cited per asset. Where no firm figure was found, the cell is blank (n/a). No estimation has been applied."),
    ("Staff Acom Jebel Ali", "Not included in this paste of the inventory (no GLA recorded). If it should be in the report, add it and re-run."),
    ("EH Jumeirah Garden", "Format not specified in the source. Assumed residential (likely Jumeirah Garden City, Al Satwa). Unit count not found under this exact name."),
    ("Brand", "Palette and typography drawn from the betterhomes CT1-Visualizer brand variables: slate, denim, powder, sand, mist, bronze."),
]
for label, text in notes:
    ws.cell(row=row, column=1, value=label).font = BODY_BOLD
    ws.cell(row=row, column=1).alignment = LEFT
    ws.cell(row=row, column=1).fill = FILL_MIST
    tc = ws.cell(row=row, column=2, value=text); tc.font = BODY; tc.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 48
    row += 1

row += 1
ws.cell(row=row, column=1, value="Sources by asset").font = H3
row += 1
header_row(ws, row, ["Asset / topic", "Finding", "URL"])
row += 1
src_rows = [
    ("Falcon Tower",                         "s_falcon"),
    ("City Tower 2",                         "s_ct2"),
    ("Green Tower",                          "s_green"),
    ("Splendour Villas",                     "s_splendour"),
    ("Al Sufouh Villas",                     "s_sufouh"),
    ("Galleria Villas Al Wasl",              "s_galleria_v"),
    ("Al Barsha Villas A / B / C",           "s_barsha_v"),
    ("EH Jumeirah Garden",                   "s_eh"),
    ("Galleria Al Wasl mall",                "s_gal_wasl"),
    ("Galleria Al Barsha mall",              "s_gal_barsha"),
    ("Millennium Tower",                     "s_millennium"),
    ("Medcare Hospital",                     "s_medcare"),
    ("Four Seasons Jumeirah Beach",          "s_fs_jb"),
    ("Four Seasons DIFC",                    "s_fs_difc"),
]
for label, key in src_rows:
    finding, url = SOURCES[key]
    ws.cell(row=row, column=1, value=label).font = BODY_BOLD
    ws.cell(row=row, column=1).alignment = LEFT
    fc = ws.cell(row=row, column=2, value=finding); fc.font = BODY; fc.alignment = LEFT
    uc = ws.cell(row=row, column=3, value=url if url else "(no single canonical URL)"); uc.font = BODY_MUTED; uc.alignment = LEFT
    if url:
        uc.hyperlink = url
        uc.font = Font(name="Calibri", size=10, color="2C537A", underline="single")
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 40
    row += 1

ws.freeze_panes = "A3"

# ---------------------------------------------------------------------------
out_path = "/home/user/CT1-Visualizer/reports/Brightstar_portfolio_report.xlsx"
wb.save(out_path)
print(f"wrote {out_path}")
